"""
Serviço para geração do Relatório de Gestão (Excel) com duas abas:
- RESULTADO_SIMBOLICO: utiliza apenas dias com liquido_pregao > 0 (ganhos) – visão otimista.
- RESULTADO_REAL: utiliza todos os dias com relatório enviado (inclui prejuízos) – visão real.
Regras aplicadas:
- Clientes isentos são EXCLUÍDOS de todos os cálculos (totais, médias, repasses).
- Clientes sem nenhuma nota no período não entram nos cálculos agregados, mas aparecem em seção separada.
- Para a aba real, repasse 30% é zerado se o total mensal for negativo.
"""

from datetime import datetime
import calendar
import pytz
from app import db
from app.models import User, Fatura, FaturaDiaria
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import tempfile
import os

tz_br = pytz.timezone('America/Sao_Paulo')

# Cores para linhas de clientes
AZUL_CLARO = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
VERDE_CLARO = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
BRANCO = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# Estilos comuns
def get_styles():
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    currency_format = '"R$" #,##0.00'
    percent_format = '0.00"%"'
    return {
        'header_fill': header_fill,
        'header_font': header_font,
        'subheader_fill': subheader_fill,
        'thin_border': thin_border,
        'center_align': center_align,
        'left_align': left_align,
        'right_align': right_align,
        'currency_format': currency_format,
        'percent_format': percent_format
    }


def gerar_relatorio_gestao(mes, ano):
    """
    Gera relatório com duas abas: RESULTADO_SIMBOLICO (apenas dias positivos)
    e RESULTADO_REAL (todos os dias, incluindo negativos).
    Retorna caminho do arquivo Excel temporário.
    """
    primeiro_dia = datetime(ano, mes, 1).date()
    ultimo_dia = datetime(ano, mes, calendar.monthrange(ano, mes)[1]).date()
    styles = get_styles()

    wb = Workbook()
    # Remove a aba padrão criada automaticamente
    wb.remove(wb.active)

    # ==================== ABA 1: RESULTADO_SIMBOLICO ====================
    ws_sim = wb.create_sheet("RESULTADO_SIMBOLICO")
    _criar_aba_relatorio(ws_sim, mes, ano, primeiro_dia, ultimo_dia, simbolico=True, styles=styles)

    # ==================== ABA 2: RESULTADO_REAL ====================
    ws_real = wb.create_sheet("RESULTADO_REAL")
    _criar_aba_relatorio(ws_real, mes, ano, primeiro_dia, ultimo_dia, simbolico=False, styles=styles)

    fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    wb.save(temp_path)
    return temp_path


def _criar_aba_relatorio(ws, mes, ano, primeiro_dia, ultimo_dia, simbolico, styles):
    """
    Cria o conteúdo de uma aba (simbólica ou real) no worksheet fornecido.
    simbolico=True -> apenas dias com liquido_pregao > 0 (e bruto > 0 para ROI)
    simbolico=False -> todos os dias com status 'relatorio_enviado'
    """
    # Buscar clientes ativos não isentos
    clientes_ativos = User.query.filter(
        User.role == 'cliente',
        User.status_acesso == 'ativo',
        User.is_isento == False
    ).all()
    clientes_ativos_qtd = len(clientes_ativos)

    # Acumuladores globais
    total_capital = 0.0
    soma_liquido_total = 0.0
    total_dias_operados = 0
    repasse_todos_clientes_total = 0.0
    repasse_comissionados_total = 0.0
    rois_clientes = []

    # Listas de clientes com e sem dados
    dados_clientes = []
    clientes_sem_notas = []

    for cliente in clientes_ativos:
        # Filtro base: dias do período com relatório enviado
        query_dias = FaturaDiaria.query.join(Fatura).filter(
            Fatura.user_id == cliente.id,
            FaturaDiaria.data_pregao >= primeiro_dia,
            FaturaDiaria.data_pregao <= ultimo_dia,
            FaturaDiaria.status == 'relatorio_enviado'
        )

        if simbolico:
            # Simbólico: apenas dias positivos (liquido_pregao > 0)
            dias_financeiros = query_dias.filter(FaturaDiaria.liquido_pregao > 0).all()
            dias_roi = query_dias.filter(FaturaDiaria.bruto > 0).all()
        else:
            # Real: todos os dias
            dias_financeiros = query_dias.all()
            dias_roi = dias_financeiros  # Para ROI usa o bruto real (pode ser negativo)

        capital_cliente = cliente.capital_alocado or 0.0
        total_capital += capital_cliente

        if not dias_financeiros:
            # Cliente sem nenhum dia (ou sem dias positivos no caso simbólico)
            clientes_sem_notas.append({
                'nome': cliente.nome,
                'cpf': cliente.cpf,
                'conta_mt5': cliente.conta_mt5 or '',
                'modelo': cliente.modelo_negocio,
                'capital': capital_cliente
            })
            # Contribui para ROI se tiver dias_roi (no caso real, dias_roi pode estar vazio; no simbólico, já tratado)
            if dias_roi and capital_cliente > 0:
                qtd_roi = len(dias_roi)
                soma_bruto = sum(d.bruto for d in dias_roi)
                media_bruta = soma_bruto / qtd_roi if qtd_roi else 0.0
                roi_cliente = (media_bruta / capital_cliente) * 100 if capital_cliente else 0.0
                rois_clientes.append(roi_cliente)
            continue

        # Cálculos financeiros
        qtd_dias = len(dias_financeiros)
        soma_liquido = sum(d.liquido_pregao for d in dias_financeiros)
        media_diaria = soma_liquido / qtd_dias if qtd_dias else 0.0
        media_semanal = media_diaria * 5
        total_mensal = soma_liquido

        # Repasse: 30% sobre total_mensal, mas se total_mensal for negativo, repasse = 0 (não se cobra de prejuízo)
        repasse_cliente = total_mensal * 0.30 if total_mensal > 0 else 0.0

        # Acumular globais
        total_dias_operados += qtd_dias
        soma_liquido_total += soma_liquido
        repasse_todos_clientes_total += repasse_cliente
        if cliente.modelo_negocio == 'comissao':
            repasse_comissionados_total += repasse_cliente

        # ROI (baseado em bruto, seguindo mesma lógica do dashboard)
        if dias_roi and capital_cliente > 0:
            qtd_roi = len(dias_roi)
            soma_bruto = sum(d.bruto for d in dias_roi)
            media_bruta = soma_bruto / qtd_roi if qtd_roi else 0.0
            roi_cliente = (media_bruta / capital_cliente) * 100 if capital_cliente else 0.0
            rois_clientes.append(roi_cliente)

        dados_clientes.append({
            'nome': cliente.nome,
            'cpf': cliente.cpf,
            'conta_mt5': cliente.conta_mt5 or '',
            'modelo': cliente.modelo_negocio,
            'capital': capital_cliente,
            'media_diaria': media_diaria,
            'media_semanal': media_semanal,
            'total_mensal': total_mensal,
            'repasse': repasse_cliente
        })

    # Cálculos globais
    media_diaria_global = soma_liquido_total / total_dias_operados if total_dias_operados > 0 else 0.0
    media_semanal_global = media_diaria_global * 5
    total_mensal_global = soma_liquido_total
    roi_global = sum(rois_clientes) / len(rois_clientes) if rois_clientes else 0.0

    # Escrever cabeçalho e totais
    titulo = f"RELATÓRIO DE GESTÃO - {'SIMBÓLICO' if simbolico else 'REAL'} - {mes}/{ano}"
    ws.merge_cells('A1:H1')
    ws['A1'] = titulo
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = styles['center_align']
    ws['A1'].fill = styles['header_fill']
    ws['A1'].font = styles['header_font']

    linha = 3
    ws.cell(row=linha, column=1, value='INDICADOR').fill = styles['subheader_fill']
    ws.cell(row=linha, column=2, value='VALOR').fill = styles['subheader_fill']
    linha += 1
    ws.cell(row=linha, column=1, value='Clientes Ativos (não isentos)')
    ws.cell(row=linha, column=2, value=clientes_ativos_qtd)
    linha += 1
    ws.cell(row=linha, column=1, value='Capital Total Alocado (R$)')
    ws.cell(row=linha, column=2, value=total_capital).number_format = styles['currency_format']
    linha += 1
    ws.cell(row=linha, column=1, value='Média Diária Global (R$)')
    ws.cell(row=linha, column=2, value=media_diaria_global).number_format = styles['currency_format']
    linha += 1
    ws.cell(row=linha, column=1, value='Média Semanal Global (R$)')
    ws.cell(row=linha, column=2, value=media_semanal_global).number_format = styles['currency_format']
    linha += 1
    ws.cell(row=linha, column=1, value='Total Mensal Líquido (R$)')
    ws.cell(row=linha, column=2, value=total_mensal_global).number_format = styles['currency_format']
    linha += 1
    ws.cell(row=linha, column=1, value='REPASSE 30% (Todos clientes) (R$)')
    ws.cell(row=linha, column=2, value=repasse_todos_clientes_total).number_format = styles['currency_format']
    linha += 1
    ws.cell(row=linha, column=1, value='REPASSE 30% (Comissionados) (R$)')
    ws.cell(row=linha, column=2, value=repasse_comissionados_total).number_format = styles['currency_format']
    linha += 1
    ws.cell(row=linha, column=1, value='ROI Médio Global (%)')
    ws.cell(row=linha, column=2, value=round(roi_global, 2)).number_format = styles['percent_format']

    # Aplicar bordas nos totais
    for row in range(3, linha+1):
        for col in [1,2]:
            cell = ws.cell(row=row, column=col)
            cell.border = styles['thin_border']
            if col == 2:
                cell.alignment = styles['right_align']
            else:
                cell.alignment = styles['left_align']

    # ==================== DETALHAMENTO POR CLIENTE ====================
    if dados_clientes:
        linha += 2
        ws.cell(row=linha, column=1, value='DETALHAMENTO POR CLIENTE ATIVO').font = Font(bold=True, size=12)
        linha += 1

        headers = [
            'Nome', 'CPF', 'Conta MT5', 'Modelo', 'Capital (R$)',
            'Média Diária (R$)', 'Média Semanal (R$)', 'Total Mensal (R$)', 'Repasse 30% (R$)'
        ]
        for idx, header in enumerate(headers):
            cell = ws.cell(row=linha, column=1+idx, value=header)
            cell.fill = styles['header_fill']
            cell.font = styles['header_font']
            cell.border = styles['thin_border']
            cell.alignment = styles['center_align']

        linha += 1
        for cli in dados_clientes:
            if cli['modelo'] == 'comissao':
                bg_fill = AZUL_CLARO
            else:
                bg_fill = VERDE_CLARO

            ws.cell(row=linha, column=1, value=cli['nome']).border = styles['thin_border']
            ws.cell(row=linha, column=2, value=cli['cpf']).border = styles['thin_border']
            ws.cell(row=linha, column=3, value=cli['conta_mt5']).border = styles['thin_border']
            modelo_str = 'Comissão' if cli['modelo'] == 'comissao' else 'Compra'
            ws.cell(row=linha, column=4, value=modelo_str).border = styles['thin_border']

            cell_capital = ws.cell(row=linha, column=5, value=round(cli['capital'], 2))
            cell_capital.number_format = styles['currency_format']
            cell_capital.border = styles['thin_border']
            cell_capital.fill = bg_fill

            cell_media_diaria = ws.cell(row=linha, column=6, value=round(cli['media_diaria'], 2))
            cell_media_diaria.number_format = styles['currency_format']
            cell_media_diaria.border = styles['thin_border']
            cell_media_diaria.fill = bg_fill

            cell_media_semanal = ws.cell(row=linha, column=7, value=round(cli['media_semanal'], 2))
            cell_media_semanal.number_format = styles['currency_format']
            cell_media_semanal.border = styles['thin_border']
            cell_media_semanal.fill = bg_fill

            cell_total = ws.cell(row=linha, column=8, value=round(cli['total_mensal'], 2))
            cell_total.number_format = styles['currency_format']
            cell_total.border = styles['thin_border']
            cell_total.fill = bg_fill

            cell_repasse = ws.cell(row=linha, column=9, value=round(cli['repasse'], 2))
            cell_repasse.number_format = styles['currency_format']
            cell_repasse.border = styles['thin_border']
            cell_repasse.fill = bg_fill

            for col in range(1, 10):
                cell = ws.cell(row=linha, column=col)
                if col >= 5:
                    cell.alignment = styles['right_align']
                else:
                    cell.alignment = styles['left_align']
                cell.fill = bg_fill

            linha += 1

    # ==================== CLIENTES SEM MOVIMENTAÇÃO ====================
    if clientes_sem_notas:
        linha += 2
        ws.cell(row=linha, column=1, value='CLIENTES ATIVOS SEM NOTAS NO PERÍODO').font = Font(bold=True, size=12)
        linha += 1

        headers_sem = ['Nome', 'CPF', 'Conta MT5', 'Modelo', 'Capital (R$)']
        for idx, header in enumerate(headers_sem):
            cell = ws.cell(row=linha, column=1+idx, value=header)
            cell.fill = styles['header_fill']
            cell.font = styles['header_font']
            cell.border = styles['thin_border']
            cell.alignment = styles['center_align']

        linha += 1
        for cli in clientes_sem_notas:
            modelo_str = 'Comissão' if cli['modelo'] == 'comissao' else 'Compra'
            ws.cell(row=linha, column=1, value=cli['nome']).border = styles['thin_border']
            ws.cell(row=linha, column=2, value=cli['cpf']).border = styles['thin_border']
            ws.cell(row=linha, column=3, value=cli['conta_mt5']).border = styles['thin_border']
            ws.cell(row=linha, column=4, value=modelo_str).border = styles['thin_border']
            cell_capital = ws.cell(row=linha, column=5, value=round(cli['capital'], 2))
            cell_capital.number_format = styles['currency_format']
            cell_capital.border = styles['thin_border']

            for col in range(1, 6):
                cell = ws.cell(row=linha, column=col)
                if col == 5:
                    cell.alignment = styles['right_align']
                else:
                    cell.alignment = styles['left_align']
            linha += 1

    # Ajustar larguras das colunas
    column_widths = [30, 15, 12, 12, 15, 15, 15, 15, 18]
    for i, width in enumerate(column_widths, start=1):
        col_letter = chr(64 + i) if i <= 26 else 'A' + chr(64 + (i-26))
        ws.column_dimensions[col_letter].width = width
