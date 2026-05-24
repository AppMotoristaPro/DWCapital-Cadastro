import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, timedelta
import pytz

tz_br = pytz.timezone('America/Sao_Paulo')

def gerar_relatorio_pendencias(usuarios_ativos, caminho_saida):
    """
    Exportador Híbrido: Calcula os dias que deveriam ter notas na vida do cliente (desde data de cadastro até ontem), 
    descarta os dias que já foram preenchidos no banco de dados e salva na planilha formatada.
    Agora com cores diferentes para clientes comissionados (azul claro) e compra de licença (verde claro).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NOTAS PENDENTES"

    # Cores para cabeçalho
    azul_fill = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")
    fonte_branca_bold = Font(color="FFFFFF", bold=True, size=12)
    fonte_preta = Font(color="000000", size=11)
    alinhamento_centro = Alignment(horizontal="center", vertical="center")
    
    # Cores para linhas de clientes
    cor_comissao = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")   # azul bem claro
    cor_compra   = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")   # verde bem claro
    cor_padrao   = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")   # branco (fallback)
    
    borda_fina = Border(
        left=Side(style='thin', color="000000"), 
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"), 
        bottom=Side(style='thin', color="000000")
    )

    ws.merge_cells('A1:B1')
    ws['A1'] = "NOTAS PENDENTES DE LANÇAMENTO"
    ws['A1'].fill = azul_fill
    ws['A1'].font = fonte_branca_bold
    ws['A1'].alignment = alinhamento_centro
    ws['A1'].border = borda_fina
    ws['B1'].border = borda_fina

    ws['A2'] = "NOME"
    ws['A2'].fill = azul_fill
    ws['A2'].font = fonte_branca_bold
    ws['A2'].alignment = alinhamento_centro
    ws['A2'].border = borda_fina

    ws['B2'] = "DATA"
    ws['B2'].fill = azul_fill
    ws['B2'].font = fonte_branca_bold
    ws['B2'].alignment = alinhamento_centro
    ws['B2'].border = borda_fina

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20

    hoje = datetime.now(tz_br).date()
    ontem = hoje - timedelta(days=1)
    
    dados = []
    
    for user in usuarios_ativos:
        data_cadastro = user.data_cadastro.date() if user.data_cadastro else datetime.min.date()
        
        # Cataloga os dias reais gravados no banco que o cliente já resolveu
        dias_resolvidos = set()
        for fatura in user.faturas:
            for dia in fatura.dias:
                if dia.status in ['relatorio_enviado', 'isento']:
                    dias_resolvidos.add(dia.data_pregao)
        
        # Gera o calendário virtual de obrigações dele
        dias_pendentes = []
        data_atual = data_cadastro
        
        # Condição crucial: Só cobra o que for menor ou igual a ONTEM
        while data_atual <= ontem:
            if data_atual.weekday() < 5: 
                if data_atual not in dias_resolvidos:
                    dias_pendentes.append(data_atual.strftime('%d/%m/%Y'))
            data_atual += timedelta(days=1)
            
        if dias_pendentes:
            # Determina a cor de acordo com o modelo_negocio
            modelo = getattr(user, 'modelo_negocio', 'comissao')
            if modelo == 'compra':
                cor_fundo = cor_compra
            else:  # comissao
                cor_fundo = cor_comissao
            
            dados.append({
                'nome': user.nome.upper(),
                'datas': dias_pendentes,
                'cor': cor_fundo
            })

    if not dados:
        return False

    linha_atual = 3

    for cliente in dados:
        nome = cliente['nome']
        datas = cliente['datas']
        cor_cliente = cliente['cor']
        
        linha_inicio = linha_atual
        
        for data in datas:
            celula_data = ws.cell(row=linha_atual, column=2, value=data)
            celula_data.alignment = alinhamento_centro
            celula_data.font = fonte_preta
            celula_data.border = borda_fina
            celula_data.fill = cor_cliente   # aplica a cor do cliente
            linha_atual += 1
            
        linha_fim = linha_atual - 1
        
        if linha_inicio < linha_fim:
            ws.merge_cells(start_row=linha_inicio, start_column=1, end_row=linha_fim, end_column=1)
        
        celula_nome = ws.cell(row=linha_inicio, column=1, value=nome)
        celula_nome.alignment = alinhamento_centro
        celula_nome.font = fonte_preta
        celula_nome.fill = cor_cliente   # aplica a cor na célula mesclada do nome
        
        for r in range(linha_inicio, linha_fim + 1):
            ws.cell(row=r, column=1).border = borda_fina
            # garantia de cor nas células vazias do nome mesclado (já foi aplicada na primeira)
            if r > linha_inicio:
                ws.cell(row=r, column=1).fill = cor_cliente

    # Adicionar legenda (opcional) nas primeiras linhas
    ws.merge_cells('D1:E1')
    ws['D1'] = "LEGENDA:"
    ws['D1'].font = Font(bold=True, size=10)
    ws['D2'] = "Azul claro:"
    ws['D2'].font = Font(size=9)
    ws['D2'].fill = cor_comissao
    ws['E2'] = "Cliente comissionado (30%)"
    ws['E2'].font = Font(size=9)
    ws['D3'] = "Verde claro:"
    ws['D3'].font = Font(size=9)
    ws['D3'].fill = cor_compra
    ws['E3'] = "Cliente com licença (compra de robô)"
    ws['E3'].font = Font(size=9)
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 25

    wb.save(caminho_saida)
    return True
